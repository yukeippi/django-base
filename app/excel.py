import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from .models import Task

HEADERS = ['タイトル', '説明', 'ステータス', '優先度', '期限']
VALID_STATUSES = {choice[0] for choice in Task.STATUS_CHOICES}


# Excelファイルを読み込んでTaskを一括作成する
# 戻り値: (作成件数, エラーリスト)
# エラーリスト要素: {'row': 行番号, 'message': エラー内容}
# 想定列順: タイトル | 説明 | ステータス | 優先度 | 期限
def import_tasks_from_excel(file, has_header=True):
    # data_only=True で数式セルを計算済みの値として読む
    # （Excelで一度保存されたファイルのみ有効。未保存の場合はNoneになる）
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    data_rows = rows[1:] if has_header else rows
    start_row = 2 if has_header else 1

    created_count = 0
    errors = []

    for offset, row in enumerate(data_rows):
        row_num = start_row + offset
        padded = list(row) + [None] * 5
        title, description, status, priority, due_date = padded[:5]

        if not title:
            errors.append({'row': row_num, 'message': 'タイトルが空です'})
            continue

        if status and status not in VALID_STATUSES:
            errors.append({'row': row_num, 'message': f'ステータスが不正です: {status}'})
            continue

        if priority is not None:
            try:
                priority = int(priority)
                if not (1 <= priority <= 5):
                    raise ValueError
            except (ValueError, TypeError):
                errors.append({'row': row_num, 'message': f'優先度は1〜5の整数で入力してください: {priority}'})
                continue

        parsed_due_date = None
        if due_date:
            if hasattr(due_date, 'date'):
                parsed_due_date = due_date.date()
            else:
                parsed_due_date = parse_date(str(due_date))
                if parsed_due_date is None:
                    errors.append({'row': row_num, 'message': f'期限の日付形式が不正です: {due_date}'})
                    continue

        Task.objects.create(
            title=str(title),
            description=str(description) if description else '',
            status=status or 'todo',
            priority=priority if priority is not None else 3,
            due_date=parsed_due_date,
        )
        created_count += 1

    return created_count, errors


# TaskのQuerySetを新規Excelファイルとしてレスポンスで返す
def export_tasks_to_new_excel(tasks):
    wb = Workbook()
    ws = wb.active
    ws.title = 'タスク'
    ws.append(HEADERS)

    for task in tasks:
        ws.append([
            task.title,
            task.description,
            task.get_status_display(),
            task.priority,
            str(task.due_date) if task.due_date else '',
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tasks.xlsx"'
    wb.save(response)
    return response


# 既存のテンプレートExcelにTaskデータを書き込んでレスポンスで返す
# テンプレートのレイアウトに合わせて DATA_START_ROW と各列番号を調整すること
def export_tasks_to_template(tasks, template_path):
    DATA_START_ROW = 2  # データ開始行（テンプレートに合わせて変更）
    col_title = 1
    col_description = 2
    col_status = 3
    col_priority = 4
    col_due_date = 5

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    for offset, task in enumerate(tasks):
        row = DATA_START_ROW + offset
        ws.cell(row=row, column=col_title, value=task.title)
        ws.cell(row=row, column=col_description, value=task.description)
        ws.cell(row=row, column=col_status, value=task.get_status_display())
        ws.cell(row=row, column=col_priority, value=task.priority)
        ws.cell(row=row, column=col_due_date, value=str(task.due_date) if task.due_date else '')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="tasks_output.xlsx"'
    wb.save(response)
    return response
