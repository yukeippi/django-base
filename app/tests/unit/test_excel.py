import io
from datetime import date

import openpyxl
import pytest

from app.excel import TaskExcelHandler
from app.models import Task


def make_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@pytest.mark.django_db
class TestTaskExcelHandlerImport:

    def test_ヘッダーありで全フィールドが正しくインポートされる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクA', '説明A', 'todo', 2, '2026-05-01'],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert errors == []
        task = Task.objects.get(title='タスクA')
        assert task.description == '説明A'
        assert task.status == 'todo'
        assert task.priority == 2
        assert task.due_date == date(2026, 5, 1)

    def test_ヘッダーなしで1行目からデータとして読み込まれる(self):
        f = make_excel([['タスクB', '', 'in_progress', 1, '']])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=False)

        assert count == 1
        assert errors == []
        assert Task.objects.get(title='タスクB').status == 'in_progress'

    def test_ステータスが空の場合はデフォルト値todoになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクC', '', '', '', ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert errors == []
        assert Task.objects.get(title='タスクC').status == 'todo'

    def test_優先度が空の場合はデフォルト値3になる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクD', '', '', None, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert errors == []
        assert Task.objects.get(title='タスクD').priority == 3

    def test_期限が空の場合はNoneになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクE', '', 'todo', 3, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert errors == []
        assert Task.objects.get(title='タスクE').due_date is None

    def test_タイトルが空の行はエラーになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['', '説明', 'todo', 3, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert errors[0]['row'] == 2
        assert 'タイトルが空です' in errors[0]['message']

    def test_不正なステータスはエラーになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクF', '', 'invalid', 3, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1

    def test_優先度が範囲外の場合はエラーになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクG', '', 'todo', 9, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1

    def test_不正な日付形式はエラーになる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['タスクH', '', 'todo', 3, 'not-a-date'],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert '日付形式が不正' in errors[0]['message']

    def test_エラーがある場合は全件ロールバックされる(self):
        f = make_excel([
            ['タイトル', '説明', 'ステータス', '優先度', '期限'],
            ['正常タスク', '', 'todo', 3, ''],
            ['', '', 'todo', 3, ''],
            ['別の正常タスク', '', 'done', 1, ''],
        ])
        count, errors = TaskExcelHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert not Task.objects.filter(title='正常タスク').exists()
        assert not Task.objects.filter(title='別の正常タスク').exists()


@pytest.mark.django_db
class TestTaskExcelHandlerExportNewExcel:

    def test_タスクなしの場合はヘッダー行のみのファイルが返る(self):
        response = TaskExcelHandler().export_to_new_excel(Task.objects.none())

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 1
        assert rows[0] == ('タイトル', '説明', 'ステータス', '優先度', '期限')

    def test_タスクのデータが正しく書き出される(self):
        Task.objects.create(
            title='出力タスク',
            description='説明文',
            status='in_progress',
            priority=2,
            due_date=date(2026, 6, 1),
        )
        response = TaskExcelHandler().export_to_new_excel(Task.objects.all())

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][0] == '出力タスク'
        assert rows[1][1] == '説明文'
        assert rows[1][2] == 'in_progress'
        assert rows[1][3] == 2
        assert rows[1][4] == '2026-06-01'

    def test_ContentTypeがExcel形式になっている(self):
        response = TaskExcelHandler().export_to_new_excel(Task.objects.none())
        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_ContentDispositionにファイル名が含まれる(self):
        response = TaskExcelHandler().export_to_new_excel(Task.objects.none())
        assert 'attachment' in response['Content-Disposition']
        assert 'tasks.xlsx' in response['Content-Disposition']


@pytest.mark.django_db
class TestTaskExcelHandlerExportTemplate:

    def test_テンプレートの指定セルにデータが正しく書き込まれる(self, tmp_path):
        template_path = tmp_path / 'template.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['タイトル', '説明', 'ステータス', '優先度', '期限'])
        wb.save(str(template_path))

        Task.objects.create(
            title='テンプレートタスク',
            description='テスト説明',
            status='done',
            priority=5,
            due_date=date(2026, 7, 1),
        )
        response = TaskExcelHandler().export_to_template(
            Task.objects.all(), str(template_path)
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 'テンプレートタスク'
        assert ws.cell(row=2, column=2).value == 'テスト説明'
        assert ws.cell(row=2, column=4).value == 5

    def test_複数タスクが連続する行に書き込まれる(self, tmp_path):
        template_path = tmp_path / 'template.xlsx'
        openpyxl.Workbook().save(str(template_path))

        Task.objects.create(title='タスク1', status='todo', priority=1)
        Task.objects.create(title='タスク2', status='done', priority=2)
        response = TaskExcelHandler().export_to_template(
            Task.objects.all().order_by('id'), str(template_path)
        )

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 'タスク1'
        assert ws.cell(row=3, column=1).value == 'タスク2'

    def test_ContentDispositionにファイル名が含まれる(self, tmp_path):
        template_path = tmp_path / 'template.xlsx'
        openpyxl.Workbook().save(str(template_path))

        response = TaskExcelHandler().export_to_template(
            Task.objects.none(), str(template_path)
        )
        assert 'tasks.xlsx' in response['Content-Disposition']
