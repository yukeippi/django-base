import io

import openpyxl
import pytest

from app.excel_base import ColumnDef, ExcelHandler
from app.models import Task


class TestColumnDef:

    def test_必須パラメータのみで作成できる(self):
        col = ColumnDef(model_field='title', excel_header='タイトル')
        assert col.model_field == 'title'
        assert col.excel_header == 'タイトル'
        assert col.required is False
        assert col.cell_to_value is None
        assert col.value_to_cell is None

    def test_全パラメータを指定できる(self):
        to_val = lambda v: v.strip() if v else ''
        to_cell = str
        col = ColumnDef(
            model_field='title',
            excel_header='タイトル',
            required=True,
            cell_to_value=to_val,
            value_to_cell=to_cell,
        )
        assert col.required is True
        assert col.cell_to_value is to_val
        assert col.value_to_cell is to_cell


# テスト用Excelファイルをメモリ上に作成するヘルパー
def make_excel(rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# Taskモデルを使った最小構成のハンドラー（汎用ロジックのテスト用）
class MinimalTaskHandler(ExcelHandler):
    model = Task
    filename = 'test.xlsx'
    columns = [
        ColumnDef(model_field='title', excel_header='タイトル', required=True),
    ]


@pytest.mark.django_db
class TestExcelHandlerImport正常系:

    def test_ヘッダーありで1件インポートできる(self):
        f = make_excel([
            ['タイトル'],
            ['テストタスク'],
        ])
        count, errors = MinimalTaskHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert errors == []
        assert Task.objects.filter(title='テストタスク').exists()

    def test_ヘッダーなしで1行目からデータとして読み込まれる(self):
        f = make_excel([['テストタスク2']])
        count, errors = MinimalTaskHandler().import_from_excel(f, has_header=False)

        assert count == 1
        assert errors == []
        assert Task.objects.filter(title='テストタスク2').exists()

    def test_複数行を一括インポートできる(self):
        f = make_excel([
            ['タイトル'],
            ['タスクA'],
            ['タスクB'],
            ['タスクC'],
        ])
        count, errors = MinimalTaskHandler().import_from_excel(f, has_header=True)

        assert count == 3
        assert errors == []

    def test_cell_to_valueで変換された値が保存される(self):
        class UpperHandler(ExcelHandler):
            model = Task
            filename = 'test.xlsx'
            columns = [
                ColumnDef(
                    model_field='title',
                    excel_header='タイトル',
                    required=True,
                    cell_to_value=lambda v: v.upper() if v else v,
                ),
            ]

        f = make_excel([['タイトル'], ['hello']])
        count, errors = UpperHandler().import_from_excel(f, has_header=True)

        assert count == 1
        assert Task.objects.filter(title='HELLO').exists()


@pytest.mark.django_db
class TestExcelHandlerImportエラー系:

    def test_requiredフィールドが空の場合はエラーになる(self):
        f = make_excel([['タイトル'], ['']])
        count, errors = MinimalTaskHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert errors[0]['row'] == 2
        assert 'タイトルが空です' in errors[0]['message']

    def test_full_cleanエラーが発生した場合はエラーリストに追加される(self):
        class InvalidStatusHandler(ExcelHandler):
            model = Task
            filename = 'test.xlsx'
            columns = [
                ColumnDef(model_field='title',  excel_header='タイトル', required=True),
                ColumnDef(model_field='status', excel_header='ステータス',
                          cell_to_value=lambda v: v or 'todo'),
            ]

        f = make_excel([
            ['タイトル', 'ステータス'],
            ['タスクA', 'invalid_status'],
        ])
        count, errors = InvalidStatusHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1

    def test_cell_to_valueがValueErrorを発生した場合はエラーになる(self):
        def strict_converter(v):
            if v == 'bad':
                raise ValueError('変換エラー')
            return v

        class StrictHandler(ExcelHandler):
            model = Task
            filename = 'test.xlsx'
            columns = [
                ColumnDef(
                    model_field='title',
                    excel_header='タイトル',
                    required=True,
                    cell_to_value=strict_converter,
                ),
            ]

        f = make_excel([['タイトル'], ['bad']])
        count, errors = StrictHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert '変換エラー' in errors[0]['message']

    def test_エラーが1件でもあれば全件ロールバックされる(self):
        f = make_excel([
            ['タイトル'],
            ['正常タスク'],
            [''],           # エラー行
            ['別の正常タスク'],
        ])
        count, errors = MinimalTaskHandler().import_from_excel(f, has_header=True)

        assert count == 0
        assert len(errors) == 1
        assert not Task.objects.filter(title='正常タスク').exists()
        assert not Task.objects.filter(title='別の正常タスク').exists()


@pytest.mark.django_db
class TestExcelHandlerExportNewExcel:

    def test_ヘッダー行が正しく出力される(self):
        response = MinimalTaskHandler().export_to_new_excel(Task.objects.none())

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[0] == ('タイトル',)

    def test_データ行が正しく出力される(self):
        Task.objects.create(title='出力テスト')
        response = MinimalTaskHandler().export_to_new_excel(Task.objects.all())

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[1][0] == '出力テスト'

    def test_value_to_cellで変換された値が出力される(self):
        class UpperExportHandler(ExcelHandler):
            model = Task
            filename = 'test.xlsx'
            columns = [
                ColumnDef(
                    model_field='title',
                    excel_header='タイトル',
                    value_to_cell=lambda v: v.upper() if v else v,
                ),
            ]

        Task.objects.create(title='hello')
        response = UpperExportHandler().export_to_new_excel(Task.objects.all())

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        rows = list(wb.active.iter_rows(values_only=True))
        assert rows[1][0] == 'HELLO'

    def test_ContentTypeがExcel形式になっている(self):
        response = MinimalTaskHandler().export_to_new_excel(Task.objects.none())
        assert response['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    def test_filenameがContentDispositionに含まれる(self):
        response = MinimalTaskHandler().export_to_new_excel(Task.objects.none())
        assert 'test.xlsx' in response['Content-Disposition']
