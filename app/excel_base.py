from dataclasses import dataclass
from typing import Callable, Optional

import openpyxl
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpResponse


@dataclass
class ColumnDef:
    model_field: str
    excel_header: str
    required: bool = False
    cell_to_value: Optional[Callable] = None
    value_to_cell: Optional[Callable] = None


class ExcelHandler:
    model = None
    # サブクラスで ColumnDef のリストを定義すること。実行時に直接 append しないこと。
    columns = []
    sheet_name = 'Sheet1'
    filename = 'export.xlsx'

    def import_from_excel(self, file, has_header=True):
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_rows = rows[1:] if has_header else rows
        start_row = 2 if has_header else 1

        instances = []
        errors = []

        for offset, row in enumerate(data_rows):
            row_num = start_row + offset
            padded = list(row) + [None] * len(self.columns)
            kwargs = {}
            row_error = None

            for i, col_def in enumerate(self.columns):
                raw_value = padded[i]

                if col_def.required and (raw_value is None or str(raw_value).strip() == ''):
                    row_error = {'row': row_num, 'message': f'{col_def.excel_header}が空です'}
                    break

                try:
                    value = col_def.cell_to_value(raw_value) if col_def.cell_to_value else raw_value
                except (ValueError, TypeError) as e:
                    row_error = {'row': row_num, 'message': str(e)}
                    break

                kwargs[col_def.model_field] = value

            if row_error:
                errors.append(row_error)
                continue

            instance = self.model(**kwargs)
            try:
                instance.full_clean()
                instances.append(instance)
            except ValidationError as e:
                errors.append({'row': row_num, 'message': str(e)})

        if errors:
            return 0, errors

        with transaction.atomic():
            for instance in instances:
                instance.save()

        return len(instances), []

    def export_to_new_excel(self, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        ws.append([col.excel_header for col in self.columns])

        for obj in queryset:
            row = []
            for col_def in self.columns:
                value = getattr(obj, col_def.model_field)
                if col_def.value_to_cell:
                    value = col_def.value_to_cell(value)
                row.append(value)
            ws.append(row)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        return response

    def export_to_template(self, queryset, template_path):
        DATA_START_ROW = 2
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active

        for offset, obj in enumerate(queryset):
            row = DATA_START_ROW + offset
            for col_idx, col_def in enumerate(self.columns, start=1):
                value = getattr(obj, col_def.model_field)
                if col_def.value_to_cell:
                    value = col_def.value_to_cell(value)
                ws.cell(row=row, column=col_idx, value=value)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}"'
        wb.save(response)
        return response
